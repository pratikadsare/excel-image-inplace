# --- Image Opening (Marketplace-based header preservation + PDF skip + dynamic filename) ---
import io
import re
import traceback
import requests
import streamlit as st
from typing import List, Set, Optional
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="Image Opening", layout="wide")

# ---------------- URL helpers ----------------
HTTP_URL_RE = re.compile(r'^(https?://)?([A-Za-z0-9\.\-]+\.[A-Za-z]{2,})(/.*)$', re.IGNORECASE)
LIKELY_CDN_RE = re.compile(r'^(cdn\.|media\.|images\.|static\.)', re.IGNORECASE)
DEFAULT_SCHEME = "https://"

def is_url_like(s: str) -> bool:
    if not isinstance(s, str): return False
    s = s.strip()
    if not s: return False
    if s.lower().startswith(("http://", "https://")): return True
    if HTTP_URL_RE.match(s): return True
    if LIKELY_CDN_RE.match(s): return True
    return False

def normalize_url(s: str, default_scheme: str = DEFAULT_SCHEME) -> Optional[str]:
    if not s: return None
    s = s.strip().strip('"').strip("'")
    if s.lower().startswith(("http://", "https://")): return s
    m = HTTP_URL_RE.match(s)
    if m:
        scheme, host, path = m.group(1), m.group(2), m.group(3)
        scheme = scheme or default_scheme
        return f"{scheme}{host}{path}"
    if LIKELY_CDN_RE.match(s):
        if "/" in s:
            host, path = s.split("/", 1)
            return f"{default_scheme}{host}/{path}"
        return f"{default_scheme}{s}"
    return None

# HEAD/GET content-type detection (for skipping PDFs and non-images)
def get_content_type(url: str, timeout: float = 10.0) -> Optional[str]:
    try:
        r = requests.head(url, timeout=timeout, allow_redirects=True)
        ct = r.headers.get("Content-Type")
        if ct: return ct.lower()
    except Exception:
        pass
    try:
        r = requests.get(url, timeout=timeout, stream=True, allow_redirects=True)
        ct = r.headers.get("Content-Type")
        if ct: return ct.lower()
    except Exception:
        pass
    return None

def is_image_content_type(ct: Optional[str]) -> bool:
    return bool(ct and ct.startswith("image/"))

# ---------------- Excel helpers ----------------
def px_to_col_width(px:int)->float: return round(px/7.0, 2)
def px_to_row_height(px:int)->float: return round(px*0.75, 2)

def detect_url_columns(ws, header_row:int=1)->List[int]:
    hits = []
    for c in range(1, ws.max_column + 1):
        for r in range(header_row + 1, min(ws.max_row, header_row + 50) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and is_url_like(v):
                hits.append(c); break
    return hits

def columns_by_names(ws, names:List[str], header_row:int=1)->List[int]:
    name_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str): name_map[v.strip()] = c
    return [name_map[n] for n in names if n in name_map]

def iter_target_cells(ws, target_cols:Set[int], header_row:int):
    for r in range(header_row + 1, ws.max_row + 1):
        for c in target_cols:
            yield ws.cell(row=r, column=c)

def adjust_dimensions(ws, col_indices:Set[int], row_height_px:int, preserve_top_rows:int):
    for c in col_indices:
        ws.column_dimensions[get_column_letter(c)].width = px_to_col_width(row_height_px)
    target_h_pt = px_to_row_height(row_height_px)
    for r in range(1, ws.max_row + 1):
        if r <= max(0, preserve_top_rows):
            continue
        ws.row_dimensions[r].height = target_h_pt

def place_anchor_image(ws, cell, url: str, w_px: int, h_px: int, keep_note: bool):
    resp = requests.get(url, timeout=25)
    resp.raise_for_status()
    data = io.BytesIO(resp.content)
    img = XLImage(data)
    img.width = w_px
    img.height = h_px
    img.anchor = cell.coordinate
    ws.add_image(img)
    if keep_note: cell.comment = Comment(f"Original URL:\\n{url}", "PreviewBot")

# ---------------- UI ----------------
st.title("Image Opening")
uploaded = st.file_uploader("Upload your Excel masterfile (.xlsx)", type=["xlsx"])

try:
    if uploaded:
        data = uploaded.read()
        wb = load_workbook(io.BytesIO(data))
        sheets = wb.sheetnames

        st.sidebar.header("Settings")

        # Marketplace + preserve-top-rows control
        marketplace = st.sidebar.selectbox(
            "Marketplace (for header row preservation)",
            ["Walmart", "Target/Mirakl", "eBay", "Amazon"],
            index=0
        )
        preserve_map = {"Walmart": 6, "Target/Mirakl": 2, "eBay": 1, "Amazon": 4}
        keep_header_heights = st.sidebar.checkbox(
            f"Keep top rows at original height (recommended for {marketplace})",
            value=True
        )
        preserve_top_rows = preserve_map[marketplace] if keep_header_heights else 0

        # Sheet selection (default All sheets)
        sheet_mode = st.sidebar.radio("Sheets to process", ["One sheet", "All sheets"], index=1, horizontal=True)
        if sheet_mode == "One sheet":
            sheet_name = st.sidebar.selectbox("Which sheet?", sheets, index=0)
            target_sheets = [sheet_name]
        else:
            target_sheets = sheets

        # Size & options
        header_row = st.sidebar.number_input("Header row number (for detecting column names)", min_value=1, value=1, step=1)
        width  = st.sidebar.number_input("Image width (px)",  min_value=40, value=140, step=10)
        height = st.sidebar.number_input("Image height (px)", min_value=40, value=140, step=10)
        keep_notes = st.sidebar.checkbox("Keep original URL as cell note", value=True)
        create_adjacent = st.sidebar.checkbox("Create preview in NEW adjacent column(s)", value=False,
                                              help="Keep URL column intact; add *_preview column to the right.")

        # Detect URL columns
        ws0 = wb[target_sheets[0]]
        headers = []
        for c in range(1, ws0.max_column + 1):
            v = ws0.cell(row=header_row, column=c).value
            headers.append(v.strip() if isinstance(v, str) else f"Col {c}")
        auto_cols_idx = detect_url_columns(ws0, header_row=header_row)
        auto_cols_names = [headers[i-1] if i-1 < len(headers) else f"Col {i}" for i in auto_cols_idx]

        selected_by_name = st.multiselect(
            "Pick URL columns by header (auto-detected suggestions pre-selected)",
            options=headers,
            default=auto_cols_names
        )

        def headers_to_indices(ws, names:List[str])->Set[int]:
            return set(columns_by_names(ws, names, header_row=header_row))

        # Preview summary
        total_urls = 0
        preview_rows = []
        for s in target_sheets:
            ws = wb[s]
            targets = headers_to_indices(ws, selected_by_name) if selected_by_name else set(detect_url_columns(ws, header_row=header_row))
            count = 0
            for cell in iter_target_cells(ws, targets, header_row=header_row):
                v = cell.value if not create_adjacent else ws.cell(row=cell.row, column=cell.column - 1).value
                if isinstance(v, str) and is_url_like(v):
                    count += 1
            total_urls += count
            preview_rows.append((s, len(targets), count))

        st.dataframe(
            {"Sheet": [r[0] for r in preview_rows],
             "Target columns": [r[1] for r in preview_rows],
             "URL cells found": [r[2] for r in preview_rows]},
            use_container_width=True
        )

        # Process images
        if st.button("Process & Prepare Download", type="primary"):
            processed = inserted = skipped_nonimage = failed = 0
            progress = st.progress(0)
            status = st.empty()

            for s in target_sheets:
                ws = wb[s]
                targets = headers_to_indices(ws, selected_by_name) if selected_by_name else set(detect_url_columns(ws, header_row=header_row))
                if not targets: continue

                # Decide preview columns (same or adjacent)
                preview_targets = set()
                if create_adjacent:
                    inserted_cols = 0
                    for col_idx in sorted(targets):
                        insert_at = col_idx + 1 + inserted_cols
                        ws.insert_cols(insert_at)
                        base = ws.cell(row=header_row, column=col_idx + inserted_cols).value
                        if isinstance(base, str): base = base.strip()
                        ws.cell(row=header_row, column=insert_at).value = f"{base or f'Col {col_idx}'}_preview"
                        preview_targets.add(insert_at)
                        inserted_cols += 1
                else:
                    preview_targets = targets

                # Resize grid (preserving top N rows)
                adjust_dimensions(ws, preview_targets, row_height_px=max(width, height), preserve_top_rows=preserve_top_rows)

                # Insert images
                for cell in iter_target_cells(ws, preview_targets, header_row=header_row):
                    src_cell = cell if not create_adjacent else ws.cell(row=cell.row, column=cell.column - 1)
                    val = src_cell.value
                    if not (isinstance(val, str) and is_url_like(val)):
                        continue
                    url = normalize_url(val) or val.strip()

                    ct = get_content_type(url)
                    if not is_image_content_type(ct):
                        skipped_nonimage += 1
                        if keep_notes and not create_adjacent:
                            try:
                                cell.comment = Comment(f"Skipped (non-image: {ct or 'unknown'})\n{url}", "PreviewBot")
                            except Exception:
                                pass
                        processed += 1
                        if total_urls:
                            progress.progress(min(processed/total_urls, 1.0))
                            status.write(f"Processed {processed}/{total_urls} | inserted:{inserted}, skipped:{skipped_nonimage}, failed:{failed}")
                        continue

                    try:
                        place_anchor_image(ws, cell, url, width, height, keep_note=keep_notes and not create_adjacent)
                        inserted += 1
                    except Exception as e:
                        failed += 1
                        if keep_notes and not create_adjacent:
                            try:
                                cell.comment = Comment(f"Preview failed; kept value.\n{url}\nError: {e}", "PreviewBot")
                            except Exception:
                                pass
                    finally:
                        processed += 1
                        if total_urls:
                            progress.progress(min(processed/total_urls, 1.0))
                            status.write(f"Processed {processed}/{total_urls} | inserted:{inserted}, skipped:{skipped_nonimage}, failed:{failed}")

            # Save file + dynamic naming
            out = io.BytesIO()
            wb.save(out); out.seek(0)
            orig_name = uploaded.name
            out_name = orig_name[:-5] + "-preview.xlsx" if orig_name.lower().endswith(".xlsx") else orig_name + "-preview.xlsx"

            status.write(f"Completed {processed}/{total_urls} | inserted:{inserted}, skipped:{skipped_nonimage}, failed:{failed}")
            st.success("Done! Download your updated workbook below.")
            st.download_button("⬇️ Download updated masterfile", data=out,
                               file_name=out_name,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Upload your .xlsx masterfile to begin.")
except Exception:
    st.error("An error occurred while running the app. See details below.")
    st.exception(traceback.format_exc())
